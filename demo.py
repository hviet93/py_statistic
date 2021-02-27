import openpyxl
import numpy

wb_obj = openpyxl.load_workbook('norris.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
norrisArr = []

# Loop will print all values
# of first column
for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    norrisArr.append(cell_obj.value)  # Tao mang so gio cua 200 bong den
    # print(cell_obj.value)

# So gio trung binh moi bong den trong 200 bong den
a = numpy.array(norrisArr)  # Nhap mang vao numpy
print(numpy.mean(a))  # Tinh gia tri trung binh cua mang
# print(norrisArr)
# cell_obj = sheet_obj.cell(row=1, column=1)
# print(cell_obj.value)
