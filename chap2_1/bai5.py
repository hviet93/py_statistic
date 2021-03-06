import openpyxl
import matplotlib.pyplot as plt


class ExportData:
    def __init__(self, file_name):
        self.data_input = {
            'brown': {
                'freq': 0,
                'per_freq': 0
            },
            'miller': {
                'freq': 0,
                'per_freq': 0
            },
            'johnson': {
                'freq': 0,
                'per_freq': 0
            },
            'jones': {
                'freq': 0,
                'per_freq': 0
            },
            'smith': {
                'freq': 0,
                'per_freq': 0
            },
            'williams': {
                'freq': 0,
                'per_freq': 0
            }
        }
        self.wb_obj = openpyxl.load_workbook('./../' + str(file_name))
        self.column_idx = 1
        self.row_idx = 2
        self.max_items = 0
        self.calc_data()
        self.calc_percent_data()

    def change_data_input(self, data_input):
        self.data_input = data_input
        self.calc_data()
        self.calc_percent_data()

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row
        self.max_items = 0

        for i in range(self.row_idx, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=self.column_idx)

            if cell_obj.value is not None:
                if isinstance(cell_obj.value, int):
                    cell_value = str(cell_obj.value)
                else:
                    cell_value = cell_obj.value.lower()
                self.max_items += 1

                if cell_value in self.data_input:
                    name_obj = self.data_input[cell_value]
                    name_obj['freq'] += 1

    def calc_percent_data(self):
        for key in self.data_input:
            data_freq = self.data_input[key]['freq']
            self.data_input[key]['per_freq'] = round(data_freq / self.max_items * 100)


def main5a():
    fig, ax = plt.subplots(1, 1)
    calc_data = ExportData('2012names.xlsx')
    data = calc_data.data_input
    column_labels = ["Name", "Relative Frequency", "Percent Frequency"]
    ax.axis('tight')
    ax.axis('off')
    table_data = []

    for key in data:
        sub_data = [str(key).capitalize(), data[key]['freq'], data[key]['per_freq']]
        table_data.append(sub_data)

    table_data.append(['Total', 50, 100])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    plt.show()


def main5b():
    calc_data = ExportData('2012names.xlsx')
    data = calc_data.data_input
    plt.style.use('ggplot')
    x = []
    name_data = []

    for key in data:
        x.append(str(key).capitalize())
        name_data.append(data[key]['freq'])

    x_pos = [i for i, _ in enumerate(x)]

    plt.bar(x_pos, name_data, color='green')
    plt.xlabel('Name')
    plt.ylabel('Frequency')

    plt.xticks(x_pos, x)
    plt.show()


def main5c():
    calc_data = ExportData('2012names.xlsx')
    data = calc_data.data_input
    fig = plt.figure()
    ax = fig.add_axes([0, 0, 1, 1])
    ax.axis('equal')
    label = []
    name_data = []

    for key in data:
        label.append(str(key).capitalize())
        name_data.append(data[key]['freq'])

    ax.pie(name_data, labels=label, autopct='%1.2f%%')
    plt.show()


# main5a()
# main5b()
# main5c()
