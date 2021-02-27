import openpyxl
import matplotlib.pyplot as plt


class TvShow:
    def __init__(self, freq):
        self.freq = freq
        self.total_item = 0

    def get_freq_percentage(self):
        return self.freq / self.total_item * 100


class ExportData:
    def __init__(self):
        self.fig, self.ax = plt.subplots(1, 1)
        self.wof = TvShow(0)
        self.thm = TvShow(0)
        self.jep = TvShow(0)
        self.jj = TvShow(0)
        self.ows = TvShow(0)
        self.wb_obj = openpyxl.load_workbook('./../syndicated.xlsx')
        self.table_data = []
        self.freq_data = []
        self.freq_percent_data = []
        self.calc_data()

    @staticmethod
    def get_data_label():
        return ['Jep', 'JJ', 'OWS', 'THM', 'WoF']

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row
        max_item = 0

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)

            if cell_obj.value is not None:
                max_item += 1

                if cell_obj.value.lower() == 'wof':
                    self.wof.freq += 1
                elif cell_obj.value.lower() == 'thm':
                    self.thm.freq += 1
                elif cell_obj.value.lower() == 'jep':
                    self.jep.freq += 1
                elif cell_obj.value.lower() == 'jj':
                    self.jj.freq += 1
                elif cell_obj.value.lower() == 'ows':
                    self.ows.freq += 1

        self.jep.total_item = self.jj.total_item = self.ows.total_item = self.thm.total_item = self.wof.total_item = max_item

        self.table_data = [
            ['Jep', self.jep.freq, int(self.jep.get_freq_percentage())],
            ['JJ', self.jj.freq, int(self.jj.get_freq_percentage())],
            ['OWS', self.ows.freq, int(self.ows.get_freq_percentage())],
            ['THM', self.thm.freq, int(self.thm.get_freq_percentage())],
            ['WoF', self.wof.freq, int(self.wof.get_freq_percentage())],
            ['Total', self.jep.freq + self.jj.freq + self.ows.freq + self.thm.freq + self.wof.freq, 100],
        ]
        self.freq_data = [self.jep.freq, self.jj.freq, self.ows.freq, self.thm.freq, self.wof.freq]
        self.freq_percent_data = [
            int(self.jep.get_freq_percentage()),
            int(self.jj.get_freq_percentage()),
            int(self.ows.get_freq_percentage()),
            int(self.thm.get_freq_percentage()),
            int(self.wof.get_freq_percentage()),
        ]

    def show_frequency_data(self):
        column_labels = ["Show", "Relative Frequency", "Percent Frequency"]
        self.ax.axis('tight')
        self.ax.axis('off')
        self.ax.table(cellText=self.table_data, colLabels=column_labels, loc="center")

        plt.show()


def main():
    show_data = ExportData()
    show_data.show_frequency_data()
