import openpyxl
import matplotlib.pyplot as plt


# Excellent, Very Good, Average, Poor, Terrible

class HotelExportData:
    def __init__(self):
        self.data_input = {
            'excellent': {
                'freq': 0,
                'per_freq': 0
            },
            'very_good': {
                'freq': 0,
                'per_freq': 0
            },
            'average': {
                'freq': 0,
                'per_freq': 0
            },
            'poor': {
                'freq': 0,
                'per_freq': 0
            },
            'terrible': {
                'freq': 0,
                'per_freq': 0
            }
        }
        self.max_items = 0
        self.wb_obj = openpyxl.load_workbook('./../hotelratings.xlsx', read_only=True, data_only=True)

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)

            if cell_obj.value is not None:
                cell_value = str(cell_obj.value).lower().strip().replace(" ", "_")
                self.max_items += 1

                if cell_value in self.data_input:
                    self.data_input[cell_value]['freq'] += 1

    def calc_percent_data(self):
        for key in self.data_input:
            data_freq = self.data_input[key]['freq']
            self.data_input[key]['per_freq'] = round(data_freq / self.max_items * 100)

    def change_to_disney_data(self):
        self.max_items = 1679
        self.data_input = {
            'excellent': {
                'freq': 807,
                'per_freq': 0
            },
            'very_good': {
                'freq': 521,
                'per_freq': 0
            },
            'average': {
                'freq': 200,
                'per_freq': 0
            },
            'poor': {
                'freq': 107,
                'per_freq': 0
            },
            'terrible': {
                'freq': 44,
                'per_freq': 0
            }
        }

        self.calc_percent_data()

    def show_table_data(self):
        fig, ax = plt.subplots(1, 1)
        column_labels = ["Name", "Frequency", "Percent Frequency"]
        ax.axis('tight')
        ax.axis('off')
        table_data = []

        for key in self.data_input:
            sub_data = [
                str(key).replace("_", " ").capitalize(),
                self.data_input[key]['freq'],
                self.data_input[key]['per_freq']
            ]
            table_data.append(sub_data)

        table_data.append(['Total', self.max_items, 100])
        ax.table(cellText=table_data, colLabels=column_labels, loc="center")

        plt.show()

    def show_bar_chart(self):
        plt.style.use('ggplot')
        labels = ['Excellent', 'Very Good', 'Average', 'Poor', 'Terrible']
        bar_data = []
        x_pos = [i for i, _ in enumerate(labels)]

        for key in self.data_input:
            bar_data.append(self.data_input[key]['per_freq'])

        plt.bar(x_pos, bar_data, color='green')
        plt.xlabel('Rating')
        plt.ylabel('Percent Frequency')
        plt.title('Ratting of the Sheraton Anaheim hotel')

        plt.xticks(x_pos, labels)
        plt.show()


def main():
    export_data = HotelExportData()
    # export_data.calc_data()
    export_data.change_to_disney_data()

    export_data.show_table_data()
    # export_data.show_bar_chart()


main()
