import openpyxl
import matplotlib.pyplot as plt


class ExportData:
    def __init__(self, data_input, file_name):
        self.data_input = data_input
        self.wb_obj = openpyxl.load_workbook('./../' + str(file_name))
        self.column_idx = 1
        self.max_items = 0
        self.calc_data()
        self.calc_percent_data()

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row
        self.max_items = m_row - 1

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=self.column_idx)

            if cell_obj.value is not None:
                cell_value = cell_obj.value.lower()

                if cell_value in self.data_input:
                    name_obj = self.data_input[cell_value]
                    name_obj['freq'] += 1

    def calc_percent_data(self):
        for key in self.data_input:
            data_freq = self.data_input[key]['freq']
            self.data_input[key]['per_freq'] = int(data_freq / self.max_items * 100)


def main():
    fig, ax = plt.subplots(1, 1)
    data = {
        'brown': {
            'freq': 0,
            'per_freq': 0
        },
        'miller': {
            'freq': 0,
            'per_freq': 0
        },
        'johnson': {
            'freq': 0,
            'per_freq': 0
        },
        'jones': {
            'freq': 0,
            'per_freq': 0
        },
        'smith': {
            'freq': 0,
            'per_freq': 0
        },
        'williams': {
            'freq': 0,
            'per_freq': 0
        }
    }

    calc_data = ExportData(data, '2012names.xlsx')
    data = calc_data.data_input
    column_labels = ["Name", "Relative Frequency", "Percent Frequency"]
    ax.axis('tight')
    ax.axis('off')
    table_data = []

    for key in data:
        sub_data = [str(key).capitalize(), data[key]['freq'], data[key]['per_freq']]
        table_data.append(sub_data)

    table_data.append(['Total', 50, 100])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    plt.show()


main()
