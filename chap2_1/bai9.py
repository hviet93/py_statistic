import openpyxl
import matplotlib.pyplot as plt


# (B) Business, (CSE) Computer Sciences and Engineering, (E) Education,
# (H) Humanities, (NSM) Naturel Sciences and Mathematics,
# (SBS) Social and Behavioral Sciences, (O) Other

class ExportTwoDegree:
    def __init__(self):
        self.data_collection = {
            'bachelor': {},
            'master': {}
        }
        self.data_input = {
            'b': {
                'freq': 0,
                'per_freq': 0
            },
            'cse': {
                'freq': 0,
                'per_freq': 0
            },
            'e': {
                'freq': 0,
                'per_freq': 0
            },
            'nsm': {
                'freq': 0,
                'per_freq': 0
            },
            'sbs': {
                'freq': 0,
                'per_freq': 0
            },
            'o': {
                'freq': 0,
                'per_freq': 0
            }
        }
        self.max_items = 0
        self.wb_obj = openpyxl.load_workbook('./../majors.xlsx', read_only=True, data_only=True)

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row

        for col in range(1, 3):
            key_degree = 'bachelor' if col == 1 else 'master'

            for i in range(2, m_row + 1):
                cell_obj = sheet_obj.cell(row=i, column=col)

                if cell_obj.value is not None:
                    cell_value = str(cell_obj.value).lower()

                    if cell_value in self.data_input:
                        self.data_input[cell_value]['freq'] += 1

                    if col == 1:
                        self.max_items += 1

            self.data_collection[key_degree] = self.data_input
            self.reset_input_data()

        for key_item in self.data_collection:
            data_item = self.data_collection[key_item]

            for key in data_item:
                data_freq = data_item[key]['freq']
                data_item[key]['per_freq'] = round(data_freq / self.max_items * 100)

            self.data_collection[key_item] = data_item

    def show_freq_table(self):
        fig, ax = plt.subplots(1, 1)
        column_labels = ["Name", "Relative Frequency", "Percent Frequency"]
        ax.axis('tight')
        ax.axis('off')
        table_data = []

        for keyItem in self.data_collection:
            data = self.data_collection[keyItem]

            for key in data:
                sub_data = [(str(key).upper()), data[key]['freq'], data[key]['per_freq']]
                table_data.append(sub_data)

            table_data.append(['Total', self.max_items, 100])
            ax.table(cellText=table_data, colLabels=column_labels, loc="center")

        plt.show()

    def reset_input_data(self):
        self.data_input = {
            'b': {
                'freq': 0,
                'per_freq': 0
            },
            'cse': {
                'freq': 0,
                'per_freq': 0
            },
            'e': {
                'freq': 0,
                'per_freq': 0
            },
            'nsm': {
                'freq': 0,
                'per_freq': 0
            },
            'sbs': {
                'freq': 0,
                'per_freq': 0
            },
            'o': {
                'freq': 0,
                'per_freq': 0
            }
        }


def main():
    export_data = ExportTwoDegree()
    export_data.calc_data()
    export_data.show_freq_table()


main()
