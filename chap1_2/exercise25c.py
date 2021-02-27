import openpyxl
import matplotlib.pyplot as plt
import numpy

wb_obj = openpyxl.load_workbook('./../shadow02.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
exchange = []
amexFreq = 0
nyseFreq = 0
otcFreq = 0
percentFrequencyArr = []

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    exchange.append(cell_obj.value)

    if str(cell_obj.value).lower() == 'amex':
        amexFreq += 1
    elif str(cell_obj.value).lower() == 'nyse':
        nyseFreq += 1
    elif str(cell_obj.value).lower() == 'otc':
        otcFreq += 1

# 25c
# Frequency of each market AMEX, NYSE and OTC
print('Frequency of each market AMEX, NYSE and OTC')
print('AMEX Frequency: ' + str(amexFreq))
print('NYSE Frequency: ' + str(nyseFreq))
print('OTC Frequency: ' + str(otcFreq))
print('--------------------------------------------')

# Percent frequency of each market
print('Percent frequency of each market')
percentFrequency = (amexFreq / len(exchange) * 100)
percentFrequencyArr.append(percentFrequency)
print('AMEX Percent Frequency: ' + str(percentFrequency))
percentFrequency = (nyseFreq / len(exchange) * 100)
percentFrequencyArr.append(percentFrequency)
print('NYSE Percent Frequency: ' + str(percentFrequency))
percentFrequency = (otcFreq / len(exchange) * 100)
percentFrequencyArr.append(percentFrequency)
print('OTC Percent Frequency: ' + str(percentFrequency))
print('--------------------------------------------')

# Frequency Bar chart
exchangeMarket = ['AMEX', 'NYSE', 'OTC']
plt.bar(exchangeMarket, percentFrequencyArr, color='green')
plt.ylabel('Frequency %')
plt.xlabel('Market')
plt.title('Exchange Market Frequency')
plt.show()
