import openpyxl
import numpy as np

wb_obj = openpyxl.load_workbook('./../shadow02.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
price_earning = []

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=5)

    if cell_obj.value is not None:
        price_earning.append(cell_obj.value)

print('Price/earning ratio')
print(np.mean(price_earning))
