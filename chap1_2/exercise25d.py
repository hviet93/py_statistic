import openpyxl
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('./../shadow02.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
intervalLabel = ['0-14.9', '15-29.9', '30-44.9', '45-59.9', '60-74.9']
interval0_14 = 0
interval15_29 = 0
interval30_44 = 0
interval45_59 = 0
interval60_74 = 0

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=6)

    if cell_obj.value is not None:
        if 0 <= cell_obj.value <= 14.0:
            interval0_14 += 1
        elif 15 <= cell_obj.value <= 29.9:
            interval15_29 += 1
        elif 30 <= cell_obj.value <= 44.9:
            interval30_44 += 1
        elif 45 <= cell_obj.value <= 59.9:
            interval45_59 += 1
        elif 60 <= cell_obj.value <= 74.9:
            interval60_74 += 1

# print(interval0_14)
# print(interval15_29)
# print(interval30_44)
# print(interval45_59)
# print(interval60_74)

plt.bar(intervalLabel,
        [interval0_14, interval15_29,
         interval30_44, interval45_59,
         interval60_74], width=1.0, color='green')
plt.ylabel('Frequency')
plt.xlabel('Interval')
plt.title('Gross Profit Margin')
plt.show()
