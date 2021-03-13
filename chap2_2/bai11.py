import openpyxl
import matplotlib.pyplot as plt


# 12-14, 15-17, 18-20, 21-23, 24-26

class ExportFrequencyClass:
    def __init__(self):
        self.wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/frequency.xlsx')
        self.data_input = {
            "12_14": {
                'freq': 0,
                'per_freq': 0
            },
            "15_17": {
                'freq': 0,
                'per_freq': 0
            },
            "18_20": {
                'freq': 0,
                'per_freq': 0
            },
            "21_23": {
                'freq': 0,
                'per_freq': 0
            },
            "24_26": {
                'freq': 0,
                'per_freq': 0
            }
        }
        self.max_items = 0

    def calc_data(self):
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)

            if cell_obj.value is not None:
                self.max_items += 1
                row_value = int(cell_obj.value)

                if 12 <= row_value <= 14:
                    self.data_input['12_14']['freq'] += 1
                elif 15 <= row_value <= 17:
                    self.data_input['15_17']['freq'] += 1
                elif 18 <= row_value <= 20:
                    self.data_input['18_20']['freq'] += 1
                elif 21 <= row_value <= 23:
                    self.data_input['21_23']['freq'] += 1
                elif 24 <= row_value <= 26:
                    self.data_input['24_26']['freq'] += 1

        for key in self.data_input:
            data_freq = self.data_input[key]['freq']
            self.data_input[key]['per_freq'] = round(data_freq / self.max_items * 100)

    def show_table_frequency(self):
        fig, ax = plt.subplots(1, 1)
        column_labels = ["Group", "Frequency", "Percent Frequency"]
        ax.axis('tight')
        ax.axis('off')
        table_data = []

        for key in self.data_input:
            sub_data = [
                str(key).replace("_", " - "),
                self.data_input[key]['freq'],
                self.data_input[key]['per_freq']
            ]
            table_data.append(sub_data)

        table_data.append(['Total', self.max_items, 100])
        ax.table(cellText=table_data, colLabels=column_labels, loc="center")

        plt.show()


def main():
    export_data = ExportFrequencyClass()
    export_data.calc_data()
    export_data.show_table_frequency()


main()
