import openpyxl
import matplotlib.pyplot as plt
import matplotlib as mpl

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/largecorp.xlsx')
# mpl.rcParams['figure.dpi'] = 600
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_corp = []
data_input = {
    '0_49': {
        'label': '0 - 49',
        'freq': 0,
        'per_freq': 0.0
    },
    '50_99': {
        'label': '50 - 99',
        'freq': 0,
        'per_freq': 0.0
    },
    '100_149': {
        'label': '100 - 149',
        'freq': 0,
        'per_freq': 0.0
    },
    '150_199': {
        'label': '150 - 199',
        'freq': 0,
        'per_freq': 0.0
    },
    '200_249': {
        'label': '200 - 249',
        'freq': 0,
        'per_freq': 0.0
    },
    '250_299': {
        'label': '250 - 299',
        'freq': 0,
        'per_freq': 0.0
    },
    '300_349': {
        'label': '300 - 349',
        'freq': 0,
        'per_freq': 0.0
    },
    '350_399': {
        'label': '350 - 399',
        'freq': 0,
        'per_freq': 0.0
    },
    '400_449': {
        'label': '400 - 449',
        'freq': 0,
        'per_freq': 0.0
    }
}
max_items = 0


# Min 43
# Max 443
# Step: 50
# Start: 0
# 0 - 49
# 50 - 99
# 100 - 149
# 150 - 199
# 200 - 249
# 250 - 299
# 300 - 349
# 350 - 399
# 400 - 449
def import_data_array():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)

        if cell_obj.value is not None:
            max_items += 1
            data_corp.append(cell_obj.value)


def calc_data_corp():
    for val in data_corp:
        if 0 <= val <= 49:
            data_input['0_49']['freq'] += 1
        elif 50 <= val <= 99:
            data_input['50_99']['freq'] += 1
        elif 100 <= val <= 149:
            data_input['100_149']['freq'] += 1
        elif 150 <= val <= 199:
            data_input['150_199']['freq'] += 1
        elif 200 <= val <= 249:
            data_input['200_249']['freq'] += 1
        elif 250 <= val <= 299:
            data_input['250_299']['freq'] += 1
        elif 300 <= val <= 349:
            data_input['300_349']['freq'] += 1
        elif 350 <= val <= 399:
            data_input['350_399']['freq'] += 1
        elif 400 <= val <= 449:
            data_input['400_449']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = freq_val / max_items


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["Revenue", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []
    total_relative = 0.0
    total_freq = 0

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        total_freq += data_input[key]['freq']
        total_relative += data_input[key]['per_freq']
        table_data.append(sub_data)

    table_data.append(['Total', total_freq, total_relative])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


def show_cumulative_table_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["Revenue", "Cumulative Frequency Distribution", "Relative Cumulative Frequency Distribution"]
    table_data = []
    total = 0

    for key in data_input:
        total += data_input[key]['freq']
        sub_data = [
            data_input[key]['label'], total, total / max_items
        ]
        table_data.append(sub_data)

    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    plt.savefig('cumulative.png', dpi=300)
    # plt.show()


def show_histogram():
    plt.style.use('ggplot')
    fig, ax = plt.subplots()
    labels = []
    data_freq = []

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green', width=1.0, edgecolor='black')
    plt.xticks(x_pos, labels)
    plt.xlabel('Revenue')
    plt.ylabel('Frequency')
    plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')

    plt.show()


import_data_array()
calc_data_corp()
# show_table_freq_data()
# show_cumulative_table_data()
show_histogram()
