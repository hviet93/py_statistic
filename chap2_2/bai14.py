import numpy as np
import matplotlib.pyplot as plt
import openpyxl


class ExportData14:
    # Number of class: 6
    # Class width: 2
    # 5-6, 7-8, 9-10, 11-12, 13-14, 15-16
    # min value: 6
    # max value: 15.8
    def __init__(self):
        self.wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/bai14.xlsx')
        self.data_input = {
            '5_6': {
                'label': 'Less or equal to 6.9',
                'freq': 0,
                'per_freq': 0
            },
            '7_8': {
                'label': 'Less or equal to 8.9',
                'freq': 0,
                'per_freq': 0
            },
            '9_10': {
                'label': 'Less or equal to 10.9',
                'freq': 0,
                'per_freq': 0
            },
            '11_12': {
                'label': 'Less or equal to 12.9',
                'freq': 0,
                'per_freq': 0
            },
            '13_14': {
                'label': 'Less or equal to 14.9',
                'freq': 0,
                'per_freq': 0
            },
            '15_16': {
                'label': 'Less or equal to 16.9',
                'freq': 0,
                'per_freq': 0
            }
        }
        self.dot_data = []
        self.data_freq = []
        self.data_cumulative = []
        self.max_items = 0
        self.table_data = []

    def calc_data(self):
        total = 0
        sheet_obj = self.wb_obj.active
        m_row = sheet_obj.max_row

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)

            if cell_obj.value is not None:
                self.max_items += 1
                cell_value = cell_obj.value

                self.dot_data.append(cell_value)

                if 5.0 <= float(cell_value) <= 6.9:
                    self.data_input['5_6']['freq'] += 1
                elif 7.0 <= float(cell_value) <= 8.9:
                    self.data_input['7_8']['freq'] += 1
                elif 9.0 <= float(cell_value) <= 10.9:
                    self.data_input['9_10']['freq'] += 1
                elif 11.0 <= float(cell_value) <= 12.9:
                    self.data_input['11_12']['freq'] += 1
                elif 13.0 <= float(cell_value) <= 14.9:
                    self.data_input['13_14']['freq'] += 1
                elif 15.0 <= float(cell_value) <= 16.9:
                    self.data_input['15_16']['freq'] += 1

        for key in self.data_input:
            freq_value = self.data_input[key]['freq']
            self.data_freq.append(freq_value)
            total += freq_value
            sub_data = [
                self.data_input[key]['label'], total,
                total / self.max_items
            ]
            self.table_data.append(sub_data)
            self.data_cumulative.append(total)

    def show_table_data(self):
        fig, ax = plt.subplots(1, 1)
        column_labels = ["Group", "Cumulative Frequency", "Cumulative Relative Frequency"]
        ax.axis('tight')
        ax.axis('off')

        ax.table(cellText=self.table_data, colLabels=column_labels, loc="center")
        plt.show()

    def show_table_freq_data(self):
        self.table_data = []

        for key in self.data_input:
            data_freq = self.data_input[key]['freq']
            self.data_input[key]['per_freq'] = round(data_freq / self.max_items * 100)

        fig, ax = plt.subplots(1, 1)
        column_labels = ["Group", "Frequency Distribution", "Percent Frequency Distribution"]
        ax.axis('tight')
        ax.axis('off')

        for key in self.data_input:
            label = self.data_input[key]['label']
            freq = self.data_input[key]['freq']
            per_freq = self.data_input[key]['per_freq']
            sub_data = [label, freq, per_freq]
            self.table_data.append(sub_data)

        self.table_data.append(['Total', self.max_items, 100])
        ax.table(cellText=self.table_data, colLabels=column_labels, loc="center")

        plt.show()

    def show_dot_diagram(self):
        # plt.scatter(self.dot_data, 'go')
        x_data = [
            6.0, 6.5, 7.0, 7.5, 8.0, 8.5, 9.0, 9.5,
            10.0, 10.5, 11.0, 11.5, 12.0, 12.5, 13.0,
            13.5, 14.0, 14.5, 15.0, 15.9
        ]

        plt.scatter(x=x_data, y=self.dot_data)
        plt.show()


def main():
    export_data = ExportData14()

    export_data.calc_data()
    # export_data.show_table_data()
    # export_data.show_dot_diagram()
    export_data.show_table_freq_data()


main()
