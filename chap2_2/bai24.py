import openpyxl

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/engineeringsalary.xlsx', data_only=True, read_only=True)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_starting = []
data_mid = []
stem_leaf_starting = {}
stem_leaf_mid = {}


def import_data_array():
    for i in range(2, m_row + 1):
        cell_obj1 = sheet_obj.cell(row=i, column=2)
        cell_obj2 = sheet_obj.cell(row=i, column=3)

        if cell_obj1.value is not None:
            data_starting.append(cell_obj1.value)

        if cell_obj2.value is not None:
            data_mid.append(cell_obj2.value)


def show_stem_leaf_chart():
    global stem_leaf_starting
    global stem_leaf_mid

    for i in range(0, len(data_mid)):
        stem_val1 = int(data_starting[i] / 10000)
        leaf_val1 = int(data_starting[i] / 1000) - stem_val1 * 10
        stem_val2 = int(data_mid[i] / 10000)
        leaf_val2 = int(data_mid[i] / 1000) - stem_val2 * 10

        if str(stem_val1) not in stem_leaf_starting:
            stem_leaf_starting[str(stem_val1)] = [leaf_val1]
        else:
            stem_leaf_starting[str(stem_val1)].append(leaf_val1)

        if str(stem_val2) not in stem_leaf_mid:
            stem_leaf_mid[str(stem_val2)] = [leaf_val2]
        else:
            stem_leaf_mid[str(stem_val2)].append(leaf_val2)

    stem_leaf_starting = stem_leaf_starting.items()
    stem_leaf_starting = sorted(stem_leaf_starting)
    stem_leaf_mid = sorted(stem_leaf_mid.items(), key=lambda x: int(x[0]))

    for i in range(0, len(stem_leaf_starting)):
        stem_leaf_str = str(stem_leaf_starting[i][0]) + " | "
        stem_leaf_starting[i][1].sort()

        for leaf_val in stem_leaf_starting[i][1]:
            stem_leaf_str = stem_leaf_str + " " + str(leaf_val)

        # print(stem_leaf_str)

    for i in range(0, len(stem_leaf_mid)):
        stem_leaf_str2 = " " + str(stem_leaf_mid[i][0]) + " | " if int(stem_leaf_mid[i][0]) < 10 \
            else str(stem_leaf_mid[i][0]) + " | "
        stem_leaf_mid[i][1].sort()

        for leaf_val in stem_leaf_mid[i][1]:
            stem_leaf_str2 = stem_leaf_str2 + " " + str(leaf_val)

        print(stem_leaf_str2)


import_data_array()
show_stem_leaf_chart()
