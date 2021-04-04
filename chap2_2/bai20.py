import openpyxl
import matplotlib.pyplot as plt
import numpy as np

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/ceotime.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_time = []
max_items = 0
data_input = {
    '11_12': {
        'label': '11 - 12',
        'freq': 0,
        'per_freq': 0.0
    },
    '13_14': {
        'label': '13 - 14',
        'freq': 0,
        'per_freq': 0.0
    },
    '15_16': {
        'label': '15 - 16',
        'freq': 0,
        'per_freq': 0.0
    },
    '17_18': {
        'label': '17 - 18',
        'freq': 0,
        'per_freq': 0.0
    },
    '19_20': {
        'label': '19 - 20',
        'freq': 0,
        'per_freq': 0.0
    },
    '21_22': {
        'label': '21 - 22',
        'freq': 0,
        'per_freq': 0.0
    },
    '23_24': {
        'label': '23 - 24',
        'freq': 0,
        'per_freq': 0.0
    }
}


def import_array_data():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)

        if cell_obj.value is not None:
            data_time.append(cell_obj.value)
            max_items += 1


def find_min_max_value():
    # 11 12
    # 13 14
    # 15 16
    # 17 18
    # 19 20
    # 21 22
    # 23 24
    # 19 20
    # 20 21
    # 21 22
    # 22 23
    print(np.max(data_time))
    print(np.min(data_time))


def calc_data_time():
    for val in data_time:
        if 11 <= val <= 12:
            data_input['11_12']['freq'] += 1
        elif 13 <= val <= 14:
            data_input['13_14']['freq'] += 1
        elif 15 <= val <= 16:
            data_input['15_16']['freq'] += 1
        elif 17 <= val <= 18:
            data_input['17_18']['freq'] += 1
        elif 19 <= val <= 20:
            data_input['19_20']['freq'] += 1
        elif 21 <= val <= 22:
            data_input['21_22']['freq'] += 1
        elif 23 <= val <= 24:
            data_input['23_24']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = freq_val / max_items


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["Time", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []
    total_relative = 0.0
    total_freq = 0

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        total_freq += data_input[key]['freq']
        total_relative += data_input[key]['per_freq']
        table_data.append(sub_data)

    table_data.append(['Total', total_freq, total_relative])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


def show_histogram():
    plt.style.use('ggplot')
    labels = []
    data_freq = []
    fig, ax = plt.subplots()

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green', width=1.0, edgecolor='black')
    plt.xticks(x_pos, labels)
    plt.xlabel('Hour per week in meetings')
    plt.ylabel('Frequency')
    # plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
    plt.tight_layout()

    plt.show()


import_array_data()
# find_min_max_value()
calc_data_time()
# show_table_freq_data()
show_histogram()
