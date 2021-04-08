import openpyxl
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure

figure(figsize=(8, 6), dpi=80)
wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/franchise.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_franchise = []
data_input = {
    '0_4999': {
        'label': '0 - 4999',
        'freq': 0,
        'per_freq': 0.0
    },
    '5000_9999': {
        'label': '5000 - 9999',
        'freq': 0,
        'per_freq': 0.0
    },
    '10000_14999': {
        'label': '10000 - 14999',
        'freq': 0,
        'per_freq': 0.0
    },
    '15000_19999': {
        'label': '15000 - 19999',
        'freq': 0,
        'per_freq': 0.0
    },
    '20000_24999': {
        'label': '20000 - 24999',
        'freq': 0,
        'per_freq': 0.0
    },
    '25000_29999': {
        'label': '25000 - 29999',
        'freq': 0,
        'per_freq': 0.0
    },
    '30000_34999': {
        'label': '30000 - 34999',
        'freq': 0,
        'per_freq': 0.0
    },
    '35000_39999': {
        'label': '35000 - 39999',
        'freq': 0,
        'per_freq': 0.0
    }
}
max_items = 0


# Max: 37496
# Min: 1431
# Step: 5000
# Start: 0
# End: 40000
# 0 - 4999
# 5000 - 9999
# 10000 - 14999
# 15000 - 19999
# 20000 - 24999
# 25000 - 29999
# 30000 - 34999
# 35000 - 39999
def import_data_array():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)

        if cell_obj.value is not None:
            max_items += 1
            data_franchise.append(cell_obj.value)


def calc_data_franchise():
    for val in data_franchise:
        if 0 <= val <= 4999:
            data_input['0_4999']['freq'] += 1
        elif 5000 <= val <= 9999:
            data_input['5000_9999']['freq'] += 1
        elif 10000 <= val <= 14999:
            data_input['10000_14999']['freq'] += 1
        elif 15000 <= val <= 19999:
            data_input['15000_19999']['freq'] += 1
        elif 20000 <= val <= 24999:
            data_input['20000_24999']['freq'] += 1
        elif 25000 <= val <= 29999:
            data_input['25000_29999']['freq'] += 1
        elif 30000 <= val <= 34999:
            data_input['30000_34999']['freq'] += 1
        elif 35000 <= val <= 39999:
            data_input['35000_39999']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = freq_val / max_items


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["US Location", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []
    total_relative = 0.0
    total_freq = 0

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        total_freq += data_input[key]['freq']
        total_relative += data_input[key]['per_freq']
        table_data.append(sub_data)

    table_data.append(['Total', total_freq, total_relative])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    # plt.savefig('frequency.png', dpi=500)
    plt.show()


def show_histogram():
    plt.style.use('ggplot')
    fig, ax = plt.subplots()
    labels = []
    data_freq = []

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green')
    plt.xticks(x_pos, labels)
    plt.xlabel('US Location')
    plt.ylabel('Frequency')
    plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
    fig.tight_layout()

    plt.show()


# b Construct histogram
# c Comment on the shape

import_data_array()
calc_data_franchise()
# show_table_freq_data()
show_histogram()
