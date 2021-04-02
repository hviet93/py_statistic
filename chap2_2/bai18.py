import openpyxl
import matplotlib.pyplot as plt

# 10 11 12
# 13 14 15
# 16 17 18
# 19 20 21
# 22 23 24
# 25 26 27
# 28 29 30

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/nbaplayerpts.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_input = {
    '10_12': {
        'label': '10 - 12',
        'freq': 0,
        'per_freq': 0.0
    },
    '13_15': {
        'label': '13 - 15',
        'freq': 0,
        'per_freq': 0.0
    },
    '16_18': {
        'label': '16 - 18',
        'freq': 0,
        'per_freq': 0.0
    },
    '19_21': {
        'label': '19 - 21',
        'freq': 0,
        'per_freq': 0.0
    },
    '22_24': {
        'label': '22 - 24',
        'freq': 0,
        'per_freq': 0.0
    },
    '25_27': {
        'label': '25 - 27',
        'freq': 0,
        'per_freq': 0.0
    },
    '28_30': {
        'label': '28 - 30',
        'freq': 0,
        'per_freq': 0.0
    }
}
max_items = 0
total = 0


def calc_data_ppg():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=3)

        if cell_obj.value is not None:
            cell_value = float(cell_obj.value)
            max_items += 1

            if 10.0 <= cell_value <= 12.9:
                data_input['10_12']['freq'] += 1
            elif 13.0 <= cell_value <= 15.9:
                data_input['13_15']['freq'] += 1
            elif 16.0 <= cell_value <= 18.9:
                data_input['16_18']['freq'] += 1
            elif 19.0 <= cell_value <= 21.9:
                data_input['19_21']['freq'] += 1
            elif 22.0 <= cell_value <= 24.9:
                data_input['22_24']['freq'] += 1
            elif 25.0 <= cell_value <= 27.9:
                data_input['25_27']['freq'] += 1
            elif 28.0 <= cell_value <= 30.0:
                data_input['28_30']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = freq_val / max_items


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["PPG", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        table_data.append(sub_data)
    table_data.append(['Total', max_items, 1.0])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


def show_cumulative_table_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["PPG", "Cumulative Frequency Distribution", "Relative Cumulative Frequency Distribution"]
    table_data = []
    global total

    for key in data_input:
        total += data_input[key]['freq']
        sub_data = [
            data_input[key]['label'], total, total / max_items
        ]
        table_data.append(sub_data)

    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


def show_histogram():
    labels = []
    data_freq = []

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green', width=1.0, edgecolor='black')
    plt.xticks(x_pos, labels)
    plt.xlabel('PPG')
    plt.ylabel('Frequency')

    plt.show()


calc_data_ppg()
# show_table_freq_data()
# show_cumulative_table_data()
show_histogram()
