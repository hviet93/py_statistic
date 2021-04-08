import openpyxl
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/marketindexes.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_index = []
max_items = 0
data_input = {
    '20.9_16.0': {
        'label': '(-20.0) - (-16.0)',
        'freq': 0,
        'per_freq': 0.0
    },
    '15.9_11.0': {
        'label': '(-15.9) - (-11.0)',
        'freq': 0,
        'per_freq': 0.0
    },
    '10.9_6.0': {
        'label': '(-10.9) - (-6.0)',
        'freq': 0,
        'per_freq': 0.0
    },
    '5.9_1.0': {
        'label': '(-5.9) - (-1.0)',
        'freq': 0,
        'per_freq': 0.0
    },
    '0.9_4.9': {
        'label': '(-0.9) - (4.9 )',
        'freq': 0,
        'per_freq': 0.0
    },
    '5.0_9.9': {
        'label': '(5.0) - (9.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '10.0_14.9': {
        'label': '(10.0) - (14.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '15.0_19.9': {
        'label': '(15.0) - (19.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '20.0_24.9': {
        'label': '(20.0) - (24.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '25.0_29.9': {
        'label': '(25.0) - (29.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '30.0_34.9': {
        'label': '(30.0) - (34.9)',
        'freq': 0,
        'per_freq': 0.0
    },
    '35.0_39.9': {
        'label': '(35.0) - (39.9)',
        'freq': 0,
        'per_freq': 0.0
    }
}


# Min -16.3
# Max 31.4
# Step 5
# Start -20
# End 40
# -20.0 - -16.0
# -15.9 - -11.0
# -10.9 - -6.0
# -5.9 - -1.0
# -0.9 - 4.9
# 5.0 - 9.9
# 10.0 - 14.9
# 15.0 - 19.9
# 20.0 - 24.9
# 25.0 - 29.9
# 30.0 - 34.9
# 35.0 - 39.9
def import_data_array():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=3)

        if cell_obj.value is not None:
            max_items += 1
            data_index.append(float(cell_obj.value))


def calc_data_index():
    for val in data_index:
        if -20.0 <= val <= -16.0:
            data_input['20.9_16.0']['freq'] += 1
        elif -15.9 <= val <= -11.0:
            data_input['15.9_11.0']['freq'] += 1
        elif -10.9 <= val <= -6.0:
            data_input['10.9_6.0']['freq'] += 1
        elif -5.9 <= val <= -1.0:
            data_input['5.9_1.0']['freq'] += 1
        elif -0.9 <= val <= 4.9:
            data_input['0.9_4.9']['freq'] += 1
        elif 5.0 <= val <= 9.9:
            data_input['5.0_9.9']['freq'] += 1
        elif 10.0 <= val <= 14.9:
            data_input['10.0_14.9']['freq'] += 1
        elif 15.0 <= val <= 19.9:
            data_input['15.0_19.9']['freq'] += 1
        elif 20.0 <= val <= 24.0:
            data_input['20.0_24.9']['freq'] += 1
        elif 25.0 <= val <= 29.9:
            data_input['25.0_29.9']['freq'] += 1
        elif 30.0 <= val <= 34.9:
            data_input['30.0_34.9']['freq'] += 1
        elif 35.0 <= val <= 39.0:
            data_input['35.0_39.9']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = round(freq_val / max_items, 2)
        # data_input[key]['per_freq'] = freq_val / max_items


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["Indexes", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []
    total_relative = 0.0
    total_freq = 0

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        total_freq += data_input[key]['freq']
        total_relative += data_input[key]['per_freq']
        table_data.append(sub_data)

    table_data.append(['Total', total_freq, round(total_relative, 2)])
    # table_data.append(['Total', total_freq, total_relative])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


import_data_array()
calc_data_index()
show_table_freq_data()
