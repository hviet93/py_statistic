import openpyxl

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/marathon.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_marathon = []
stem_leaf_marathon = {}


def import_data_array():
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)

        if cell_obj.value is not None:
            data_marathon.append(cell_obj.value)


def show_stem_leaf_char():
    global stem_leaf_marathon

    for val in data_marathon:
        stem_val = int(val / 10)
        leaf_val = val - stem_val * 10

        if str(stem_val) not in stem_leaf_marathon:
            stem_leaf_marathon[str(stem_val)] = [leaf_val]
        else:
            stem_leaf_marathon[str(stem_val)].append(leaf_val)

    stem_leaf_marathon = sorted(stem_leaf_marathon.items(), key=lambda x: int(x[0]))

    for stem_leaf in stem_leaf_marathon:
        stem_leaf_str = ' ' + str(stem_leaf[0]) + ' | ' if int(stem_leaf[0]) < 10 \
            else str(stem_leaf[0]) + ' | '
        stem_leaf_stretched_str = stem_leaf_str[:]
        stem_leaf_stretched_length = len(stem_leaf_stretched_str)
        stem_leaf[1].sort()

        for leaf_val in stem_leaf[1]:
            if leaf_val < 5:
                stem_leaf_str = stem_leaf_str + ' ' + str(leaf_val)
            else:
                stem_leaf_stretched_str = stem_leaf_stretched_str + ' ' + str(leaf_val)

        print(stem_leaf_str)

        if len(stem_leaf_stretched_str) > stem_leaf_stretched_length:
            print(stem_leaf_stretched_str)


import_data_array()
show_stem_leaf_char()
