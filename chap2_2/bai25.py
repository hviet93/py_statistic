import openpyxl
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/bestpayingdegrees.xlsx')
sheet_obj = wb_obj['Data']
m_row = sheet_obj.max_row
data_increase = []
max_items = 0
data_input = {
    '40_50': {
        'label': '40 - 50',
        'freq': 0,
        'per_freq': 0.0
    },
    '50_60': {
        'label': '50 - 60',
        'freq': 0,
        'per_freq': 0.0
    },
    '60_70': {
        'label': '60 - 70',
        'freq': 0,
        'per_freq': 0.0
    },
    '70_80': {
        'label': '70 - 80',
        'freq': 0,
        'per_freq': 0.0
    },
    '80_90': {
        'label': '80 - 90',
        'freq': 0,
        'per_freq': 0.0
    },
    '90_100': {
        'label': '90 - 100',
        'freq': 0,
        'per_freq': 0.0
    },
    '100_110': {
        'label': '100 - 110',
        'freq': 0,
        'per_freq': 0.0
    },
    '110_120': {
        'label': '110 - 120',
        'freq': 0,
        'per_freq': 0.0
    }
}
stem_leaf_increase = {}


# Max: 113
# Min: 43
# Step: 10
# 40 - 50
# 50 - 60
# 60 - 70
# 70 - 80
# 80 - 90
# 90 - 100
# 100 - 110
# 110 - 120
def import_data_array():
    global max_items

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=4)

        if cell_obj.value is not None:
            max_items += 1
            data_increase.append(int(round(cell_obj.value, 0)))


def calc_data_increase():
    for val in data_increase:
        if 40 <= val <= 50:
            data_input['40_50']['freq'] += 1
        elif 50 < val <= 60:
            data_input['50_60']['freq'] += 1
        elif 60 < val <= 70:
            data_input['60_70']['freq'] += 1
        elif 70 < val <= 80:
            data_input['70_80']['freq'] += 1
        elif 80 < val <= 90:
            data_input['80_90']['freq'] += 1
        elif 90 < val <= 100:
            data_input['90_100']['freq'] += 1
        elif 100 < val <= 110:
            data_input['100_110']['freq'] += 1
        elif 110 < val <= 120:
            data_input['110_120']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']

        data_input[key]['per_freq'] = freq_val / max_items


def show_histogram():
    plt.style.use('ggplot')
    fig, ax = plt.subplots()
    labels = []
    data_freq = []

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green', width=1.0, edgecolor='black')
    plt.xticks(x_pos, labels)
    plt.xlabel('Increase')
    plt.ylabel('Frequency')
    plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
    fig.tight_layout()

    plt.show()


def show_stem_leaf_char():
    global stem_leaf_increase

    for val in data_increase:
        stem_val = int(val / 10)
        leaf_val = val - stem_val * 10

        if str(stem_val) not in stem_leaf_increase:
            stem_leaf_increase[str(stem_val)] = [leaf_val]
        else:
            stem_leaf_increase[str(stem_val)].append(leaf_val)

    stem_leaf_increase = sorted(stem_leaf_increase.items(), key=lambda x: int(x[0]))

    for stem_leaf in stem_leaf_increase:
        stem_leaf_str = ' ' + str(stem_leaf[0]) + ' | ' if int(stem_leaf[0]) < 10 \
            else str(stem_leaf[0]) + ' | '
        stem_leaf[1].sort()

        for leaf_val in stem_leaf[1]:
            stem_leaf_str = stem_leaf_str + ' ' + str(leaf_val)

        print(stem_leaf_str)


import_data_array()
calc_data_increase()
# show_histogram()
show_stem_leaf_char()
