import openpyxl
import matplotlib.pyplot as plt
import numpy as np

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/ports.xlsx', read_only=True, data_only=True)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data_handle = []
max_item = 0
data_input = {
    '25_49.9': {
        'label': '25 - 49.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '50_74.9': {
        'label': '50 - 74.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '75_99.9': {
        'label': '75 - 99.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '100_124.9': {
        'label': '100 - 124.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '125_149.9': {
        'label': '100 - 149.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '150_174.9': {
        'label': '150 - 174.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '175_199.9': {
        'label': '175 - 199.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '200_224.9': {
        'label': '200 - 224.9',
        'freq': 0,
        'per_freq': 0.0
    },
    '225_250': {
        'label': '225 - 250',
        'freq': 0,
        'per_freq': 0.0
    },
}


def import_array_data():
    global max_item

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)

        if cell_obj.value is not None:
            data_handle.append(float(cell_obj.value))
            max_item += 1


def find_min_max_value():
    # 25 - 49.9
    # 50 - 74.9
    # 75 - 99.9
    # 100 - 124.9
    # 125 - 149.9
    # 150 - 174.9
    # 175 - 199.9
    # 200 - 224.9
    # 225 - 250
    print(np.max(data_handle))
    print(np.min(data_handle))


def calc_data_handle():
    for val in data_handle:
        if 25.0 <= val <= 49.9:
            data_input['25_49.9']['freq'] += 1
        elif 50.0 <= val <= 74.9:
            data_input['50_74.9']['freq'] += 1
        elif 75.0 <= val <= 99.9:
            data_input['75_99.9']['freq'] += 1
        elif 100.0 <= val <= 124.9:
            data_input['100_124.9']['freq'] += 1
        elif 125.0 <= val <= 149.9:
            data_input['125_149.9']['freq'] += 1
        elif 150.0 <= val <= 174.9:
            data_input['150_174.9']['freq'] += 1
        elif 175.0 <= val <= 199.9:
            data_input['175_199.9']['freq'] += 1
        elif 200.0 <= val <= 224.9:
            data_input['200_224.9']['freq'] += 1
        elif 225.0 <= val <= 250.9:
            data_input['225_250']['freq'] += 1

    for key in data_input:
        freq_val = data_input[key]['freq']
        data_input[key]['per_freq'] = freq_val / max_item

    print(data_input)


def show_table_freq_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ["Tons handle", "Frequency Distribution", "Relative Frequency Distribution"]
    table_data = []

    for key in data_input:
        sub_data = [
            data_input[key]['label'], data_input[key]['freq'],
            data_input[key]['per_freq']
        ]
        table_data.append(sub_data)
    table_data.append(['Total', max_item, 1.0])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")

    plt.show()


def show_histogram():
    plt.style.use('ggplot')
    labels = []
    data_freq = []
    fig, ax = plt.subplots()

    for key in data_input:
        labels.append(data_input[key]['label'])
        data_freq.append(data_input[key]['freq'])

    x_pos = [i for i, _ in enumerate(labels)]
    plt.bar(x_pos, data_freq, color='green', width=1.0, edgecolor='black')
    plt.xticks(x_pos, labels)
    plt.xlabel('Frequency')
    plt.ylabel('Tons Handle')
    plt.setp(ax.get_xticklabels(), rotation=30, horizontalalignment='right')
    plt.tight_layout()

    plt.show()


import_array_data()
# find_min_max_value()
calc_data_handle()
# show_table_freq_data()
show_histogram()
