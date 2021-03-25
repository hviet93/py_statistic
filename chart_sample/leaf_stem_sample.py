import openpyxl
import numpy as np


def main():
    wb_obj = openpyxl.load_workbook('./../data_sample/data_2_2/aptest.xlsx')
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    stem_data = []
    stem_leaf_data = {}

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)

        if cell_obj.value is not None:
            cell_value = cell_obj.value
            stem_value = int(cell_value / 10)

            if stem_value not in stem_data:
                stem_data.append(stem_value)

    stem_data = np.sort(stem_data)[::1]

    for stem in stem_data:
        stem_leaf_data[str(stem)] = []

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)

        if cell_obj.value is not None:
            cell_value = cell_obj.value
            stem_value = int(cell_value / 10)

            if str(stem_value) in stem_leaf_data:
                min_stem = stem_value * 10
                max_stem = stem_value * 10 + 9

                if min_stem <= cell_value <= max_stem:
                    leaf_value = cell_value - stem_value * 10
                    stem_leaf_data[str(stem_value)].append(leaf_value)

    for stem_leaf in stem_leaf_data:
        stem_leaf_str = " " + str(stem_leaf) + " | " \
            if int(stem_leaf) < 10 else str(stem_leaf) + " | "

        for leaf_val in stem_leaf_data[stem_leaf]:
            stem_leaf_str = stem_leaf_str + str(leaf_val) + " "

        print(stem_leaf_str)


main()
