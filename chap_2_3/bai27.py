import openpyxl
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_3/crosstab.xlsx')
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data = []
data_input = {
    'A': {
        '1_freq': 0,
        '2_freq': 0,
        '1_per_freq': 0,
        '2_per_freq': 0,
    },
    'B': {
        '1_freq': 0,
        '2_freq': 0,
        '1_per_freq': 0,
        '2_per_freq': 0,
    },
    'C': {
        '1_freq': 0,
        '2_freq': 0,
        '1_per_freq': 0,
        '2_per_freq': 0,
    }
}
max_items = 0


def import_data_array():
    global max_items

    for i in range(2, m_row + 1):
        cell_x_obj = sheet_obj.cell(row=i, column=2)
        cell_y_obj = sheet_obj.cell(row=i, column=3)

        if cell_x_obj and cell_y_obj is not None:
            max_items += 1
            cell_value = [str(cell_x_obj.value), int(cell_y_obj.value)]
            data.append(cell_value)


def calc_data():
    for val in data:
        lbl_row = val[0]
        lbl_col = str(val[1]) + '_freq'
        data_input[lbl_row][lbl_col] += 1


def show_crosstab_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ['x', '1', '2', 'Total']
    table_data = []
    total_col_1 = 0
    total_col_2 = 0

    for key in data_input:
        freq_1 = data_input[key]['1_freq']
        freq_2 = data_input[key]['2_freq']
        total_row = freq_1 + freq_2
        sub_data = [key, freq_1, freq_2, total_row]

        table_data.append(sub_data)
        total_col_1 += freq_1
        total_col_2 += freq_2

    table_data.append([
        'Total', total_col_1, total_col_2,
        total_col_1 + total_col_2
    ])

    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    # plt.savefig('frequency.png', dpi=500)
    plt.show()


def row_percentage_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ['', '1', '2', 'Total']
    table_data = []

    for key in data_input:
        freq_1 = data_input[key]['1_freq']
        freq_2 = data_input[key]['2_freq']
        freq_per_1 = round((freq_1 / (freq_1 + freq_2)) * 100, 1)
        freq_per_2 = round((freq_2 / (freq_1 + freq_2)) * 100, 1)
        sub_data = [key, freq_per_1, freq_per_2, freq_per_1 + freq_per_2]
        table_data.append(sub_data)

    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    # plt.savefig('frequency.png', dpi=500)
    plt.show()


def col_percentage_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ['', '1', '2']
    total_col_1 = 0
    total_col_2 = 0
    table_data = []

    for key in data_input:
        freq_1 = data_input[key]['1_freq']
        freq_2 = data_input[key]['2_freq']
        total_col_1 += freq_1
        total_col_2 += freq_2

    for key in data_input:
        freq_1 = data_input[key]['1_freq']
        freq_2 = data_input[key]['2_freq']
        freq_per_1 = round((freq_1 / total_col_1) * 100, 1)
        freq_per_2 = round((freq_2 / total_col_2) * 100, 1)
        sub_data = [key, freq_per_1, freq_per_2]
        table_data.append(sub_data)

    table_data.append(['Total', 100.0, 100.0])
    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    # plt.savefig('frequency.png', dpi=500)
    plt.show()


import_data_array()
calc_data()
# show_crosstab_data()
# row_percentage_data()
col_percentage_data()
