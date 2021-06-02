import openpyxl
import matplotlib.pyplot as plt

wb_obj = openpyxl.load_workbook('./../data_sample/data_2_3/crosstab2.xlsx', data_only=True, read_only=True)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
data = []
# x:
# min: 13, max 84, step: 20
# 10-29, 30-49, 50-69, 70-89
# y:
# min: 21, max: 99, step: 20
# 20-39, 40-59, 60-79, 80-99
# --------------------------
# col1: 20_29, 30_39, 0, 0
# col2: 0, 40_49, 59_60, 0
# col3: 0, 0, 50_59, 70_79
# col4: 0, 0, 0, 80_89
table_input_data = {
    '20_39': {
        '10_29': 0,
        '30_49': 0,
        '50_69': 0,
        '70_90': 0
    },
    '40_59': {
        '10_29': 0,
        '30_49': 0,
        '50_69': 0,
        '70_90': 0
    },
    '60_79': {
        '10_29': 0,
        '30_49': 0,
        '50_69': 0,
        '70_90': 0
    },
    '80_100': {
        '10_29': 0,
        '30_49': 0,
        '50_69': 0,
        '70_90': 0
    }
}


def import_data_array():
    for i in range(2, m_row + 1):
        cell_x_obj = sheet_obj.cell(row=i, column=2)
        cell_y_obj = sheet_obj.cell(row=i, column=3)

        if cell_x_obj.value and cell_y_obj.value is not None:
            cell_value = [cell_x_obj.value, cell_y_obj.value]
            data.append(cell_value)


def calc_data():
    for val in data:
        lbl_row = val[0]
        lbl_col = val[1]

        # Table input data
        for key_col in table_input_data:
            key_col_val_arr = str(key_col).split('_')
            key_col_val1 = int(key_col_val_arr[0])
            key_col_val2 = int(key_col_val_arr[1])

            if key_col_val1 <= lbl_col <= key_col_val2:
                for key_row in table_input_data[key_col]:
                    key_row_val_arr = str(key_row).split('_')
                    key_row_val1 = int(key_row_val_arr[0])
                    key_row_val2 = int(key_row_val_arr[1])

                    if key_row_val1 <= lbl_row <= key_row_val2:
                        table_input_data[key_col][key_row] += 1
                        break


def show_crosstab_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ['', '20-39', '40-59', '60-79', '80-99', 'Total']
    row_labels = ['10-29', '30-49', '50-69', '70-89']
    table_data = []
    list_key_col1 = list(table_input_data['20_39'])
    list_key_col2 = list(table_input_data['40_59'])
    list_key_col3 = list(table_input_data['60_79'])
    list_key_col4 = list(table_input_data['80_100'])
    total_col1 = 0
    total_col2 = 0
    total_col3 = 0
    total_col4 = 0

    for i in range(0, 4):
        key_col1 = list_key_col1[i]
        key_col2 = list_key_col2[i]
        key_col3 = list_key_col3[i]
        key_col4 = list_key_col4[i]
        total = table_input_data['20_39'][key_col1] + table_input_data['40_59'][key_col2] \
                + table_input_data['60_79'][key_col3] + table_input_data['80_100'][key_col4]
        total_col1 += table_input_data['20_39'][key_col1]
        total_col2 += table_input_data['40_59'][key_col2]
        total_col3 += table_input_data['60_79'][key_col3]
        total_col4 += table_input_data['80_100'][key_col4]
        sub_data = [
            row_labels[i], table_input_data['20_39'][key_col1],
            table_input_data['40_59'][key_col2], table_input_data['60_79'][key_col3],
            table_input_data['80_100'][key_col4], total
        ]
        table_data.append(sub_data)
    table_data.append([
        'Total', total_col1, total_col2,
        total_col3, total_col4, total_col1 + total_col2 + total_col3 + total_col4
    ])

    ax.table(cellText=table_data, colLabels=column_labels, loc="center")
    # plt.savefig('frequency.png', dpi=500)
    plt.show()


def show_row_percentage_data():
    fig, ax = plt.subplots(1, 1)
    ax.axis('tight')
    ax.axis('off')
    column_labels = ['', '20-39', '40-59', '60-79', '80-99', 'Total']
    row_labels = ['10-29', '30-49', '50-69', '70-89']
    list_key_col = list(table_input_data['20_39'])
    table_data = []

    for key_row in table_input_data:
        total_row = 0
        sub_data = []
        # Count total
        for i in range(0, 4):
            key_col = list_key_col[i]
            total_row = table_input_data[key_row][key_col] + table_input_data[key_row][key_col] \
                + table_input_data[key_row][key_col] + table_input_data[key_row][key_col]
            print(total_row)

        # Add data
        # for i in range(0, 4):
        #     key_col = list_key_col[i]
        #     per_row = (table_input_data[key_row][key_col] / total_row) * 100




import_data_array()
calc_data()
# show_crosstab_data()
show_row_percentage_data()
